import math
import numpy as np
import cartopy.feature as cfeature
import cartopy.io.shapereader as shpreader
import xarray as xr
import cartopy.crs as ccrs
from shapely.ops import unary_union
import matplotlib.pyplot as plt
from cartopy.mpl.gridliner import LONGITUDE_FORMATTER, LATITUDE_FORMATTER
import matplotlib.ticker as mticker
from scipy.stats import circstd, circmean, circvar
import os
import openpyxl
import pandas as pd
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from matplotlib.projections import get_projection_class


def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points
    on the earth (specified in decimal degrees)
    :param x: an array containing [lon1,lat1,lon2,lat2]
    :return: a scaler distance in km
    """
    # convert decimal degrees to radians
    lon1, lat1, lon, lat = map(math.radians, [lon1, lat1, lon2, lat2])
    # haversine formula
    dlon = lon - lon1
    dlat = lat - lat1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat) * math.sin(dlon / 2) ** 2
    c = 2 * math.asin(math.sqrt(a))
    km = 6371 * c
    return km

def initial_bearing(long1, lat1, long2, lat2):
    """
    :param long1: longitude of point 1 (in degrees)
    :param lat1:  latitude of point 1 (in degrees)
    :param long2: longitude of point 2 (in degrees)
    :param lat2: latitude of point 2 (in degrees)
    :return:
    The initial bearing on a sphere from point 1 to point 2.
    """
    theta = np.arctan2(np.sin(np.radians(long2 - long1)) * np.cos(np.radians(lat2)),
                       np.cos(np.radians(lat1)) * np.sin(np.radians(lat2)) -
                       np.sin(np.radians(lat1)) * np.cos(np.radians(lat2)) * np.cos(np.radians(long2 - long1)))
    return np.degrees(theta) % 360

def rosePlot(ax, strikes):
    bin_edges = np.arange(-5, 366, 10)
    number_of_strikes, bin_edges = np.histogram(strikes, bin_edges)
    number_of_strikes[0] += number_of_strikes[-1]
    half = np.sum(np.split(number_of_strikes[:-1], 2), 0)
    two_halves = np.concatenate([half, half])
    ax.bar(np.deg2rad(bin_edges[:-2]+5), two_halves,
               width=np.deg2rad(10), bottom=0.0, color='.8', edgecolor='k', lw=1)
    ax.set_theta_zero_location('N')
    ax.set_theta_direction(-1)
    ax.set_thetagrids(np.arange(0, 360, 10), labels=np.arange(0, 360, 10))
    ax.set_rgrids(np.arange(1, two_halves.max() + 1, 2), angle=0, weight='black')

def set_Mfigure(rect, alpha=0.4, df=None, dic=dict(c='salmon',s=8)):
    fname = "/Users/guy.bendor/Natural Earth Map/map_files/ME_Shaded6.nc"
    water_file = "/Users/guy.bendor/Natural Earth Map/map_files/SomeLayers4Guy/WaterBodiesSubset.shp"
    f2_file = "/Users/guy.bendor/Natural Earth Map/map_files/SomeLayers4Guy/Quaternary_faults_2020_update_limited_fields.shp"
    da = xr.open_dataset(fname)
    faults = cfeature.ShapelyFeature(shpreader.Reader(f2_file).geometries(), ccrs.PlateCarree())
    water = cfeature.ShapelyFeature(shpreader.Reader(water_file).geometries(), ccrs.PlateCarree())
    water = unary_union(list(water.geometries())).geoms
    da = da.sel(x=slice(rect[0], rect[1]), y=slice(rect[3], rect[2]))

    if df is not None:
        unique_ev = df.drop_duplicates('evid')

    fig = plt.figure(figsize=(10, 9.5))
    ax = plt.axes(projection=ccrs.PlateCarree())
    ax.set_extent(rect)

    ax.imshow(da.__xarray_dataarray_variable__.data[0], transform=ccrs.PlateCarree(), origin='upper',
              cmap="Greys_r", extent=rect, zorder=0, alpha=alpha)
    # da.__xarray_dataarray_variable__.plot(transform=ccrs.PlateCarree(),cmap="Greys_r", zorder=0, alpha=0.6)
    ax.add_feature(faults, facecolor='None', edgecolor='k', linestyle='-', zorder=1, label="Fault",lw=1,alpha=0.5)
    # ax.add_feature(f[1], facecolor='None', edgecolor='k', linestyle='--', zorder=1, label="Fault",lw=1,alpha=0.9)

    ax.add_geometries(water, facecolor='lightcyan', edgecolor='black', linestyle='-',
                      linewidth=0.5, crs=ccrs.PlateCarree(), zorder=0)

    gl = ax.gridlines(crs=ccrs.PlateCarree(), linewidth=0.5, color='black', alpha=0.3, linestyle='--', draw_labels=True)
    gl.top_labels = True
    gl.left_labels = True
    gl.right_labels = True
    gl.xlocator = mticker.FixedLocator(np.arange(np.floor(rect[0]), rect[1] + 2, 2))
    gl.ylocator = mticker.FixedLocator(np.arange(np.floor(rect[2]), rect[3] + 2, 2))
    gl.xformatter = LONGITUDE_FORMATTER
    gl.yformatter = LATITUDE_FORMATTER
    if df is not None:
        ax.scatter(unique_ev["lon"], unique_ev["lat"], alpha=0.8, linewidths=0.5, edgecolors='k',
                   zorder=3, label="Event",**dic)
    return fig, ax

def add_rose(fig, ax, df, plot_specs,filepath, min_num=0, max_var=90, max_std=90,arrows={}):
    # area = df.region.unique()[0]
    st_order = plot_specs['st_order']
    ds = plot_specs['ds']
    st_row = plot_specs['start_row']
    bin_edges = np.arange(-5, 366, 10)
    df["rad"] = df[f'angle_s'].apply(np.radians)
    strikeDat = df[['station', "rad"]].groupby("station").agg(
        lambda x: [circmean(x, low=-np.pi/2,high=np.pi/2), circstd(x, low=-np.pi/2,high=np.pi/2), circvar(x, low=-np.pi/2,high=np.pi/2)]
    )
    strikeDat[["meanRad", "stdRad", "varRad"]] = np.array(strikeDat.rad.tolist())
    strikeDat.drop(columns="rad", inplace=True)
    strikeDat['x'] = np.sin(strikeDat.meanRad.values)
    strikeDat['y'] = np.cos(strikeDat.meanRad.values)
    strikeDat['meanAngle'] = strikeDat.meanRad.apply(np.degrees)
    strikeDat['stdAngle'] = strikeDat.stdRad.apply(np.degrees)
    strikeDat['varAngle'] = strikeDat.varRad.apply(np.degrees)
    strikeDat = strikeDat.merge(
        df[['station', "angle_s"]].groupby("station").agg(max_n_bin),
        left_on="station",
        right_index=True, how="left"
    )
    strikeDat = strikeDat.merge(
        df[['station', "epi_dist"]].groupby("station").agg(np.max),
        left_on="station",
        right_index=True, how="left"
    )
    strikeDat = strikeDat.merge(
        df.groupby("station").size().rename('event_count'),
        left_on="station",
        right_index=True, how="left"
    )

    strikeDat = strikeDat.merge(
        df[["station", "longitude", "latitude","total_num"]].drop_duplicates(subset=["station"]), on="station", how="left"
    )
    # print(df[['station','dt']])
    strikeDat["percentage"] = strikeDat.event_count.astype(str)+'/'+strikeDat.total_num.astype(str)
    svDf = strikeDat[["station", "meanAngle", "stdAngle", "varAngle", "epi_dist", 'percentage']].round(1).sort_values("stdAngle")
    svDf = svDf.merge(
        df[['station', "dt"]].groupby("station").agg(np.mean),
        left_on="station",
        right_index=True, how="left"
    )
    print(svDf[["station", "varAngle",'dt', 'percentage']])
    # svDf.to_csv('all_stations.csv',index=False)
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        wb.save(filepath)
    # else:
    #     wb = openpyxl.load_workbook(filepath)
    #
    # if area in wb.sheetnames:
    #     print(wb.sheetnames)
    #     del wb[area]
    #     print(wb.sheetnames)

    with pd.ExcelWriter(filepath, mode='a',if_sheet_exists='replace') as writer:
        svDf.to_excel(writer, sheet_name='all_stations', index=False)



    stations = []
    for r in st_order:
        stations+=r
    strikeDat = strikeDat.set_index('station').reindex(stations).reset_index()
    # print(strikeDat)
    strikeDat['color'] = np.where(
        # (strikeDat.angle_s >= min_num) &
        (strikeDat.event_count >= min_num) &
        (strikeDat.varAngle <= max_var) &
        (strikeDat.stdAngle <= max_std), 'tab:blue', 'tab:red'#royalblue', 'red'
    )

    # print(strikeDat)

    stations = strikeDat.station.values

    # print(s)
    # dax = []
    # print(len(stations)//4,len(stations)%4)

    # d = 0.1
    axes = []
    for i in range(len(st_order)):
        if i==0:
            x = -0.02
            y = st_row[0]
        elif i==1:
            x = st_row[1]
            y = 1.03
        elif i == 2:
            x = 1.02
            y = 1 - st_row[2]
        elif i == 3:
            x = 1 - st_row[3]
            y = -0.03
        for j in range(len(st_order[i])):
            ax_sub = inset_axes(ax, width=0.9, height=0.9, loc=10,
                            bbox_to_anchor=(x,y),
                            bbox_transform=ax.transAxes,
                            borderpad=0.0, axes_class=get_projection_class("polar"))
            ax_sub.get_xaxis().set_visible(False)
            ax_sub.get_yaxis().set_visible(False)
            ax_sub.set_theta_zero_location('N')
            ax_sub.set_theta_direction(-1)
            axes.append(ax_sub)

            if i==0:
                y+=ds[i]#*1.2
            elif i==1:
                x+=ds[i]
            elif i==2:
                y-=ds[i]
            elif i==3:
                x-=ds[i]

    for i in range(len(strikeDat)):
        sta = strikeDat.iloc[i]
        temp = df[df.station == sta.station]
        strikes = temp.angle_s.values
        number_of_strikes, bin_edges = np.histogram(strikes, bin_edges)
        number_of_strikes[0] += number_of_strikes[-1]
        half = np.sum(np.split(number_of_strikes[:-1], 2), 0)
        two_halves = np.concatenate([half, half])
        binned_dir = two_halves.max()
        axes[i].bar(np.deg2rad(bin_edges[:-2] + 5), two_halves,
                                  width=np.deg2rad(10), bottom=0.0, color=sta.color, edgecolor='k', lw=1.3, zorder=1,
                                  alpha=1)  # ,zorder=2) #color='.8'


        if sta.station in st_order[0]:
            axes[i].set_title(sta.station, fontdict=dict(fontsize=8, fontweight='bold'), rotation=35, y=0.85, x=0.1,
                              va='center')
        elif sta.station in st_order[1]:
            axes[i].set_title(sta.station, fontdict=dict(fontsize=8, fontweight='bold'), va='center')
        elif sta.station in st_order[2]:
            axes[i].set_title(sta.station, fontdict=dict(fontsize=8, fontweight='bold')
                              ,rotation=-35,y=0.85,x=0.85,va='center')
        elif sta.station in st_order[3]:
            axes[i].set_title(sta.station, fontdict=dict(fontsize=8, fontweight='bold'), va='center',y=0.97)
        # axes[i].set_title(sta.station, fontdict=dict(fontsize=8, fontweight='bold'), y=0.8)



    ax.scatter(strikeDat["longitude"], strikeDat["latitude"], transform=ccrs.PlateCarree(), s=150, zorder=5,
               label='Station',
               color="yellow",
               marker='^', linewidths=0.5, edgecolors='k')
    props = dict(boxstyle='round', facecolor='white', alpha=0.7)
    props1 = dict(boxstyle='round', facecolor='palegreen', alpha=0.7)
    for i in range(len(strikeDat)):
        row = strikeDat.iloc[i]
        x, y, name,  p = row.longitude, row.latitude, row.station, props
        ha='left'
        if name in ['HDNS']:
            x += 0.001
            y -= 0.02
            p = props1
        elif name in ['DSHN','RSPN','NATI','GEM','LVBS','KDZV','KSH0','ENGV','HULT','KNHM']:
            x += 0.007
            y += 0.001
        elif name in ['RMOT','GLHS']:
            x += 0.001
            y -= 0.02
        elif name in ['MSAM','TVR','AMID','MGDL','SPR','ZEFT','KRSH','ALMT','MNHM']:
            x -= 0.01
            y -= 0.01
            ha = 'right'
        elif name in ['GOSH']:
            x -= 0.005
            y -= 0.02
            ha = 'right'
        # else:
        #     x += 0.007
        #     y += 0.001

        ax.text(x, y, name, transform=ax.transData, fontsize=7, zorder=7,
                va='bottom', bbox=p, weight='bold',ha=ha)

    strikeDat = strikeDat[

        ((strikeDat.event_count >= min_num) &
        (strikeDat.varAngle <= max_var) &
        (strikeDat.stdAngle <= max_std)) | (strikeDat.station.isin(['HDNS']))
    ]
    dat = strikeDat[['longitude', 'latitude', 'x', 'y']].values.T
    ax.quiver(dat[0], dat[1], dat[2], dat[3], zorder=6, color='k', lw=1, ec='k', **arrows)
    ax.quiver(dat[0], dat[1], -dat[2], -dat[3], zorder=6, color='k', lw=1, ec='k', **arrows)
    # scale_bar(ax, length=20, location=(0.85, 0.13), linewidth=1.5)
    return ax

def max_n_bin(strikes):
    bin_edges = np.arange(-5, 366, 10)
    number_of_strikes, bin_edges = np.histogram(strikes, bin_edges)
    number_of_strikes[0] += number_of_strikes[-1]
    half = np.sum(np.split(number_of_strikes[:-1], 2), 0)
    two_halves = np.concatenate([half, half])
    return two_halves.max()